#!/usr/bin/env python3
"""Generate aggregated players database from Excel sources.

Reads seasonal Excel files from multiple sources and produces a single
`players_database.json` with average, min and max prices for each player
grouped by role.
"""
from pathlib import Path
import json
from openpyxl import load_workbook

SOURCES = {
    "2025_26": {
        "fantaboom": "fantaboom_2025_26.xlsx",
        "fantaclassic": "fantaclassic_2025_26.xlsx",
        "profeta": "profeta_2025_26.xlsx",
        "sos_fanta": "sos_fanta_2025_26.xlsx",
    },
    "2024_25": {
        "fantaboom": "fantaboom_2024_25.xlsx",
        "fantaclassic": "fantaclassic_2024_25.xlsx",
        "profeta": "profeta_2024_25.xlsx",
        "sos_fanta": "sos_fanta_2024_25.xlsx",
    },
}


def load_source(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).lower() for c in next(rows)]
    except StopIteration:  # empty sheet
        return []
    mapping = {
        "nome": "name",
        "ruolo": "role",
        "squadra": "team",
        "prezzo": "price",
        "gol": "goals",
        "goal": "goals",
        "assist": "assists",
        "ass": "assists",
        "minuti": "minutes",
        "min": "minutes",
        "media": "rating",
        "mv": "rating",
        "commento": "comm",
    }
    cols = [mapping.get(c, c) for c in header]
    result: list[dict] = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        data = {cols[i]: row[i] for i in range(len(cols))}
        for col in ["team", "goals", "assists", "minutes", "rating", "comm"]:
            if col not in data or data[col] is None:
                data[col] = "" if col in ("team", "comm") else 0
        result.append(
            {
                "name": data.get("name", ""),
                "team": data.get("team", ""),
                "role": data.get("role", ""),
                "price": data.get("price", 0),
                "goals": data.get("goals", 0),
                "assists": data.get("assists", 0),
                "minutes": data.get("minutes", 0),
                "rating": data.get("rating", 0),
                "comm": data.get("comm", ""),
            }
        )
    return result


def main() -> None:
    frames: list[dict] = []
    for season, sources in SOURCES.items():
        for source, filename in sources.items():
            path = Path(filename)
            if not path.exists():
                continue
            for row in load_source(path):
                row["source"] = source
                row["season"] = season
                frames.append(row)

    if not frames:
        raise SystemExit("no source data found")

    data_by_player: dict[tuple[str, str], list[dict]] = {}
    for row in frames:
        key = (row["name"], row["role"])
        data_by_player.setdefault(key, []).append(row)

    result: dict[str, list[dict]] = {}
    for (name, role), grp in data_by_player.items():
        prices = [float(r.get("price", 0) or 0) for r in grp]
        avg = round(sum(prices) / len(prices), 1)
        min_p = int(min(prices))
        max_p = int(max(prices))
        player = {
            "nome": name,
            "team": grp[0].get("team", ""),
            "prezzi": {"min": min_p, "max": max_p, "avg": avg},
            "allPrices": {},
            "performance": {},
            "notes": {"comm": ""},
        }
        fallback_comm = ""
        for row in grp:
            year = str(row["season"]).split("_")[0]
            player["allPrices"][f"{row['source']}_{year}"] = int(row.get("price", 0) or 0)
            player["performance"].setdefault(
                row["season"],
                {
                    "goals": int(row.get("goals", 0) or 0),
                    "assists": int(row.get("assists", 0) or 0),
                    "minutes": int(row.get("minutes", 0) or 0),
                    "rating": float(row.get("rating", 0) or 0),
                },
            )
            comm = row.get("comm")
            if row["source"] == "sos_fanta" and isinstance(comm, str) and comm:
                player["notes"]["comm"] = comm
            elif not fallback_comm and isinstance(comm, str) and comm:
                fallback_comm = comm
        if not player["notes"]["comm"] and fallback_comm:
            player["notes"]["comm"] = fallback_comm

        for season in SOURCES.keys():
            player["performance"].setdefault(
                season,
                {"goals": 0, "assists": 0, "minutes": 0, "rating": 0},
            )

        result.setdefault(role, []).append(player)

    with Path("players_database.json").open("w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":  # pragma: no cover
    main()
